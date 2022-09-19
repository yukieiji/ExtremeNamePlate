import os
import json
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))

NAMEPLATE_IN_FILE = os.path.join(WORKING_DIR, 'NamePlateTransData.xlsx')
NAMEPLATE_FOLDER = os.path.join(WORKING_DIR, 'namePlate')
NAMEPLATE_OUT_FILE = os.path.join(WORKING_DIR, 'namePlate', 'namePlateTransData.json')
NAMEPLATE_DATA_FILE = os.path.join(WORKING_DIR, 'namePlate', 'namePlateData.json')

NAMEPLATE_LICENSE = 'LICENSE.md'
NAMEPLATE_DATA_DATA_KEY = 'data'
NAMEPLATE_DATA_UPDATE_COMIT_KEY = 'updateComitHash'

def stringToJson(filename : str, outputFile : str) -> None:

  wb = load_workbook(filename, read_only = True)

  stringData = {}
  for s in wb:
    rows = s.iter_rows(min_col = 1, min_row = 2, max_col = 17, max_row = None)
    headers = []
    for header in s[1]:
      if header.value:
        headers.append(header.value)

    for row in rows:
      name = row[0].value

      if not name:
        continue

      data = {}

      for i, string in enumerate(row[1:]):
        if string.value:
          # I hate excel why did I do this to myself
          data[i] = string.value.replace("\r", "").replace("_x000D_", "").replace("\\n", "\n")

      if data:
        stringData[name] = data

  with open(outputFile, "w") as f:
    json.dump(stringData, f, indent=4)

def update_visor_data()-> None:

  visor_dir = os.listdir(NAMEPLATE_FOLDER)

  new_visor = {}
  new_visor[NAMEPLATE_DATA_UPDATE_COMIT_KEY] = ""

  for visor in visor_dir:

    if os.path.isdir(os.path.join(NAMEPLATE_FOLDER, visor)):

      visors = os.listdir(os.path.join(NAMEPLATE_FOLDER, visor))
      cleaned_visor = [add_visor.replace('.png', '') for add_visor in visors if add_visor != NAMEPLATE_LICENSE]
      new_visor[visor] = cleaned_visor


  with open(os.path.join(NAMEPLATE_DATA_FILE), mode='w') as visor_json:
    json.dump(
      new_visor,
      visor_json, indent=2, ensure_ascii=False)

def main()-> None:
  update_visor_data()
  stringToJson(NAMEPLATE_IN_FILE, NAMEPLATE_OUT_FILE)

if __name__ == "__main__":
  main()